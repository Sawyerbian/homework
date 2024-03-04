import random
import openpyxl

def make_sheet(name):
  # 创建工作簿
  wb = openpyxl.Workbook()

  # 创建工作表
  ws = wb.active

  # 设置标题
  ws['A1'] = '日期          '
  ws['B1'] = '时间          '

  # 生成题目
  for i in range(1, 26):
    for j in range(1, 5):
      if i % 2 == 1:
        # 加法
        a = random.randint(11, 30)
        b = random.randint(11, 30)
        while(a % 10 + b % 10) // 10 < 1:
          a = random.randint(11, 30)
          b = random.randint(11, 30)
          if (a % 10 + b % 10) // 10 >= 1:
            break
        ws.cell(row=i+1, column=j).value = f"{a} + {b} =   "
      else:
        # 减法
        a = random.randint(1, 30)
        b = random.randint(5, 30)
        while((a % 10 - b % 10) >= 0 or (a < b)):
            a = random.randint(11, 30)
            b = random.randint(8, 20)
            if (a % 10 - b % 10) < 0 and (a > b):
              break
        ws.cell(row=i+1, column=j).value = f"{a} - {b} =   "
        
  for i in range(1, 5):
    if i % 2 == 1:
      # 加法
      a = random.randint(100, 200)
      b = random.randint(50, 100)
      while(a % 10 + b % 10) // 10 < 1:
        a = random.randint(100, 200)
        b = random.randint(50, 100)
        if (a % 10 + b % 10) // 10 >= 1:
          break
      ws.cell(row=27, column=i).value = f"{a} + {b} =   "

    else:
      # 减法
      a = random.randint(100, 200)
      b = random.randint(50, 100)
      while((a % 10 - b % 10) >= 0 or (a < b)):
          a = random.randint(100, 200)
          b = random.randint(50, 100)
          if (a % 10 - b % 10) < 0 and (a > b):
            break
      ws.cell(row=27, column=i).value = f"{a} - {b} =   "
  wb.save('{}.xlsx'.format(name))
  

for i in range(0,9):
  make_sheet(str(i))