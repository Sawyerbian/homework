import random
import openpyxl
from openpyxl.styles import Font
fontStyle = Font(size = "14")

def make_sheet(name):
  # 创建工作簿
  wb = openpyxl.load_workbook('计算.xlsx')

  # 创建工作表
  ws = wb.create_sheet(str(name))

  # 设置标题
  ws['A1'] = '日期          '
  ws['B1'] = '时间          '

  # 生成题目
  for i in range(1, 26):
    for j in range(1, 5):
      if i % 2 == 1:
        # 加法
        a = random.randint(5, 20)
        b = random.randint(8, 15)
        while(a % 10 + b % 10) // 10 < 1:
          a = random.randint(5, 20)
          b = random.randint(8, 15)
          if (a % 10 + b % 10) // 10 >= 1:
            break
        ws.cell(row=i+1, column=j).value = f"{a} + {b} =   "
        ws.cell(row=i+1, column=j).font = fontStyle
      else:
        # 减法
        a = random.randint(11, 20)
        b = random.randint(5, 20)
        while((a % 10 - b % 10) >= 0 or (a < b)):
            a = random.randint(11, 20)
            b = random.randint(5, 20)
            if (a % 10 - b % 10) < 0 and (a > b):
              break
        ws.cell(row=i+1, column=j).value = f"{a} - {b} =   "
        ws.cell(row=i+1, column=j).font = fontStyle

    ws.row_dimensions[i].height=23
  for col in ['A','B','C','D']:
    ws.column_dimensions[col].width=22 

  # for i in range(1, 5):
  #   if i % 2 == 1:
  #     # 加法
  #     a = random.randint(100, 200)
  #     b = random.randint(50, 100)
  #     while(a % 10 + b % 10) // 10 < 1:
  #       a = random.randint(100, 200)
  #       b = random.randint(50, 100)
  #       if (a % 10 + b % 10) // 10 >= 1:
  #         break
  #     ws.cell(row=27, column=i).value = f"{a} + {b} =   "

  #   else:
  #     # 减法
  #     a = random.randint(100, 200)
  #     b = random.randint(50, 100)
  #     while((a % 10 - b % 10) >= 0 or (a < b)):
  #         a = random.randint(100, 200)
  #         b = random.randint(50, 100)
  #         if (a % 10 - b % 10) < 0 and (a > b):
  #           break
  #     ws.cell(row=27, column=i).value = f"{a} - {b} =   "
  wb.save('{}.xlsx'.format('计算'))
  

for i in range(0,30):
  make_sheet(i)